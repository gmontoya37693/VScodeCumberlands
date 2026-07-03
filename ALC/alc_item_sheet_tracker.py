#!/usr/bin/env python3
"""
ALC Item Sheet Tracker

Objective
---------
Track leased assets for Acento Leasing Company, recompute installment values
when bank rates change, and produce daily and monthly operational reports.

Builder
-------
German Montoya - Data Scientist & AI Specialist

For
---
Acento Leasing Company (ALC)

What the script does
--------------------
- Tracks each asset by property/allocation
- Recalculates the monthly payment using bank APR + NIM converted to an effective monthly lease rate
- Produces a point-in-time portfolio snapshot
- Produces a monthly invoice list
- Produces a monthly bank-payable summary
- Supports reconciliation between asset balances and bank debt

Inputs
------
1) assets.csv
    Required columns:
    asset_id,property_id,allocation,asset_description,asset_value,start_date,
    term_months,bank_rate_annual,nim_annual,gl_account,status

    Optional columns:
    tax_amount,admin_expense,risk_cost_recovery,salvage_value,salvage_periods

2) rates.csv
    Required columns:
    effective_date,bank_rate_annual

3) As-of date or target month passed through the command line

Outputs
-------
- Daily portfolio report
- Asset-level snapshot with balance and amortization progress
- Asset-level projected amortization schedule
- Monthly invoice lines
- Monthly bank-payable summary
- CSV exports when requested

Suggested operation
-------------------
1) Run the script daily.
2) If a new asset, rate change, invoice run, or bank payment is due, follow
    the command prompt and complete the action.
3) If nothing changed, generate the daily report and exit.
4) Run month-end invoicing once per month for all active assets.

Date format: YYYY-MM-DD
Rate format: decimal annual APR (5% = 0.05)
"""

from __future__ import annotations

import argparse
import calendar
import csv
import hashlib
import json
import math
import shutil
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple


BASE_DIR = Path(__file__).resolve().parent
SCRIPT_VERSION = "2026.07.10-compliance-v1"
DEFAULT_ASSETS = BASE_DIR / "assets.csv"
DEFAULT_RATES = BASE_DIR / "rates.csv"
DEFAULT_POSTED_LEDGER = BASE_DIR / "posted_invoices.csv"
DEFAULT_CLOSED_PERIODS = BASE_DIR / "closed_periods.csv"
DEFAULT_BASELINE_CONFIG = BASE_DIR / "baseline_config.json"
DEFAULT_MANIFEST_DIR = BASE_DIR / "run_manifests"
DEFAULT_BACKUP_DIR = BASE_DIR / "backups"
DEFAULT_ONE_PAGER = BASE_DIR / "ALC - Asset Calculation Unit.xlsx"
DEFAULT_BANK_PAYABLE = BASE_DIR / "bank_payable.csv"
POSTED_LEDGER_FIELDS = [
    "asset_id",
    "property_id",
    "allocation",
    "gl_account",
    "payment_month",
    "due_date",
    "billing_date",
    "period",
    "term_to_maturity",
    "bank_rate_annual",
    "nim_annual",
    "lease_rate_monthly",
    "lease_base",
    "opening_balance",
    "payment",
    "interest",
    "amortization",
    "ending_balance",
    "rate_source",
    "posted_timestamp",
]
BANK_PAYABLE_FIELDS = [
    "month",
    "billing_date",
    "invoice_lines",
    "bank_payable_total",
    "operator",
    "updated_at",
]


# ---------------------------
# Date and Finance Utilities
# ---------------------------
def parse_date(value: str) -> date:
    """Parse common date formats used in CSV exports into a date object."""
    raw = value.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unsupported date format: {value!r}")


def month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    return date(y, m, 1)


def add_months_keep_day(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, min(d.day, last_day))


def next_working_day(d: date) -> date:
    out = d
    while out.weekday() >= 5:
        out += timedelta(days=1)
    return out


def billing_run_date_for_month(target_month: date, billing_day: int) -> date:
    month = month_start(target_month)
    if billing_day < 1 or billing_day > 31:
        raise ValueError("billing_day must be between 1 and 31")
    last_day = calendar.monthrange(month.year, month.month)[1]
    anchor = date(month.year, month.month, min(billing_day, last_day))
    return next_working_day(anchor)


def billing_cycle_window(target_month: date, billing_day: int) -> Tuple[date, date]:
    current_billing = billing_run_date_for_month(target_month, billing_day)
    previous_billing = billing_run_date_for_month(add_months(target_month, -1), billing_day)
    return previous_billing, current_billing


def months_between(start_month: date, end_month: date) -> int:
    return (end_month.year - start_month.year) * 12 + (end_month.month - start_month.month)


def pmt(principal: float, monthly_rate: float, n_periods: int) -> float:
    """Return the fixed payment amount for an annuity-style amortization."""
    if n_periods <= 0:
        return 0.0
    if abs(monthly_rate) < 1e-12:
        return principal / n_periods
    return principal * (monthly_rate / (1.0 - math.pow(1.0 + monthly_rate, -n_periods)))


def effective_monthly_rate(annual_rate: float) -> float:
    """Convert annual APR to an effective monthly rate."""
    if abs(annual_rate) < 1e-12:
        return 0.0
    return math.pow(1.0 + annual_rate, 1.0 / 12.0) - 1.0


def month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def generate_run_id() -> str:
    return f"{datetime.now().strftime('%Y%m%dT%H%M%S')}-{uuid.uuid4().hex[:8]}"


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def hash_file(path: Path) -> str:
    if not path.exists():
        return ""

    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def collect_hashes(paths: Dict[str, Path]) -> Dict[str, str]:
    return {k: hash_file(v) for k, v in paths.items()}


def backup_file(path: Path, backup_dir: Path, run_id: str) -> Optional[str]:
    if not path.exists():
        return None

    dest = backup_dir / run_id / path.name
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, dest)
    return str(dest)


def write_manifest(manifest_dir: Path, run_id: str, payload: Dict[str, object]) -> Path:
    manifest_path = manifest_dir / f"{run_id}.json"
    ensure_parent(manifest_path)
    manifest_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return manifest_path


def load_baseline_config(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_baseline_config(path: Path, cfg: Dict[str, object]) -> None:
    ensure_parent(path)
    path.write_text(json.dumps(cfg, indent=2, sort_keys=True), encoding="utf-8")


def load_closed_periods(path: Path) -> List[str]:
    if not path.exists():
        return []

    months: List[str] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            m = (row.get("month") or "").strip()
            if m:
                months.append(m)
    return sorted(set(months))


def save_closed_periods(path: Path, months: List[str]) -> None:
    ensure_parent(path)
    unique_months = sorted(set(months))
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["month"])
        writer.writeheader()
        for m in unique_months:
            writer.writerow({"month": m})


def initialize_posted_ledger(path: Path) -> None:
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=POSTED_LEDGER_FIELDS)
        writer.writeheader()


# ------------------
# Core Data Models
# ------------------
@dataclass
class Asset:
    asset_id: str
    property_id: str
    allocation: str
    asset_description: str
    asset_value: float
    start_date: date
    term_months: int
    bank_rate_annual: float
    nim_annual: float
    gl_account: str
    status: str
    tax_amount: float = 0.0
    admin_expense: float = 0.0
    risk_cost_recovery: float = 0.0
    salvage_value: float = 0.0
    salvage_periods: int = 0

    @property
    def start_month(self) -> date:
        return month_start(self.start_date)

    @property
    def final_month(self) -> date:
        return add_months_keep_day(self.start_date, self.payment_months)

    @property
    def payment_months(self) -> int:
        return max(0, self.term_months - self.salvage_periods)

    @property
    def lease_base(self) -> float:
        return max(
            0.0,
            self.asset_value + self.tax_amount + self.admin_expense + self.risk_cost_recovery - self.salvage_value,
        )


def load_assets(path: Path) -> List[Asset]:
    # Asset rows define the contract inputs used to build lease base and schedule.
    out: List[Asset] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required = {
            "asset_id",
            "property_id",
            "allocation",
            "asset_description",
            "asset_value",
            "start_date",
            "term_months",
            "bank_rate_annual",
            "nim_annual",
            "gl_account",
            "status",
        }
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"assets.csv missing columns: {sorted(missing)}")

        for row in reader:
            out.append(
                Asset(
                    asset_id=row["asset_id"].strip(),
                    property_id=row["property_id"].strip(),
                    allocation=row["allocation"].strip(),
                    asset_description=row["asset_description"].strip(),
                    asset_value=float(row["asset_value"]),
                    start_date=parse_date(row["start_date"]),
                    term_months=int(row["term_months"]),
                    bank_rate_annual=float(row["bank_rate_annual"]),
                    nim_annual=float(row["nim_annual"]),
                    gl_account=row["gl_account"].strip(),
                    status=row["status"].strip().lower(),
                    tax_amount=float(row.get("tax_amount", "0") or 0),
                    admin_expense=float(row.get("admin_expense", "0") or 0),
                    risk_cost_recovery=float(row.get("risk_cost_recovery", "0") or 0),
                    salvage_value=float(row.get("salvage_value", "0") or 0),
                    salvage_periods=int(row.get("salvage_periods", "0") or 0),
                )
            )
    return out


# -------------------------
# Input/Output Data Access
# -------------------------
def load_rates(path: Path) -> List[Tuple[date, float]]:
    # Rates are sorted so month lookups can take the latest effective rate <= target month.
    rates: List[Tuple[date, float]] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required = {"effective_date", "bank_rate_annual"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"rates.csv missing columns: {sorted(missing)}")

        for row in reader:
            rates.append((month_start(parse_date(row["effective_date"])), float(row["bank_rate_annual"])))

    rates.sort(key=lambda x: x[0])
    return rates


def load_posted_invoices(path: Path) -> List[Dict[str, object]]:
    # Posted rows are historical actuals and should override recalculated values.
    if not path.exists():
        return []

    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        out: List[Dict[str, object]] = []
        for row in reader:
            due_date = (row.get("due_date") or row.get("payment_month") or "").strip()
            billing_date = (row.get("billing_date") or "").strip()
            out.append(
                {
                    "asset_id": row["asset_id"].strip(),
                    "property_id": row.get("property_id", "").strip(),
                    "allocation": row.get("allocation", "").strip(),
                    "gl_account": row.get("gl_account", "").strip(),
                    "payment_month": due_date,
                    "due_date": due_date,
                    "billing_date": billing_date,
                    "period": int(row["period"]),
                    "term_to_maturity": int(row["term_to_maturity"]),
                    "bank_rate_annual": float(row["bank_rate_annual"]),
                    "nim_annual": float(row["nim_annual"]),
                    "lease_rate_monthly": float(row["lease_rate_monthly"]),
                    "lease_base": float(row["lease_base"]),
                    "opening_balance": float(row["opening_balance"]),
                    "payment": float(row["payment"]),
                    "interest": float(row["interest"]),
                    "amortization": float(row["amortization"]),
                    "ending_balance": float(row["ending_balance"]),
                    "rate_source": row.get("rate_source", "posted_ledger"),
                    "posted_timestamp": row.get("posted_timestamp", ""),
                }
            )
    return out


def posted_invoice_map(rows: List[Dict[str, object]]) -> Dict[Tuple[str, str], Dict[str, object]]:
    # Key by (asset_id, payment_month) for fast row replacement while building schedules.
    return {(str(r["asset_id"]), str(r["payment_month"])): r for r in rows}


def save_posted_invoices(path: Path, rows: List[Dict[str, object]]) -> None:
    # Normalize numeric precision before writing to keep ledger exports stable.
    if not rows:
        return

    normalized = []
    for row in rows:
        normalized.append(
            {
                "asset_id": row["asset_id"],
                "property_id": row["property_id"],
                "allocation": row["allocation"],
                "gl_account": row.get("gl_account", ""),
                "payment_month": row["payment_month"],
                "due_date": row.get("due_date", row["payment_month"]),
                "billing_date": row.get("billing_date", ""),
                "period": int(row["period"]),
                "term_to_maturity": int(row["term_to_maturity"]),
                "bank_rate_annual": round(float(row["bank_rate_annual"]), 6),
                "nim_annual": round(float(row["nim_annual"]), 6),
                "lease_rate_monthly": round(float(row["lease_rate_monthly"]), 8),
                "lease_base": round(float(row["lease_base"]), 2),
                "opening_balance": round(float(row["opening_balance"]), 2),
                "payment": round(float(row["payment"]), 2),
                "interest": round(float(row["interest"]), 2),
                "amortization": round(float(row["amortization"]), 2),
                "ending_balance": round(float(row["ending_balance"]), 2),
                "rate_source": row.get("rate_source", "posted_ledger"),
                "posted_timestamp": row.get("posted_timestamp", datetime.now().isoformat(timespec="seconds")),
            }
        )

    normalized.sort(key=lambda r: (str(r["asset_id"]), str(r["payment_month"])))
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=POSTED_LEDGER_FIELDS)
        writer.writeheader()
        writer.writerows(normalized)


def merge_posted_invoices(existing: List[Dict[str, object]], new_rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    # New postings replace matching month keys to keep one final row per asset/month.
    merged = posted_invoice_map(existing)
    for row in new_rows:
        merged[(str(row["asset_id"]), str(row["payment_month"]))] = row
    return list(merged.values())


def load_bank_payable_rows(path: Path) -> List[Dict[str, object]]:
    if not path.exists():
        return []

    out: List[Dict[str, object]] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            m = (row.get("month") or "").strip()
            if not m:
                continue
            out.append(
                {
                    "month": m,
                    "billing_date": (row.get("billing_date") or "").strip(),
                    "invoice_lines": int(row.get("invoice_lines") or 0),
                    "bank_payable_total": float(row.get("bank_payable_total") or 0.0),
                    "operator": (row.get("operator") or "").strip(),
                    "updated_at": (row.get("updated_at") or "").strip(),
                }
            )
    return out


def save_bank_payable_rows(path: Path, rows: List[Dict[str, object]]) -> None:
    ensure_parent(path)
    normalized = []
    for row in rows:
        normalized.append(
            {
                "month": str(row["month"]),
                "billing_date": str(row.get("billing_date", "")),
                "invoice_lines": int(row.get("invoice_lines", 0)),
                "bank_payable_total": round(float(row.get("bank_payable_total", 0.0)), 2),
                "operator": str(row.get("operator", "")),
                "updated_at": str(row.get("updated_at", "")),
            }
        )

    normalized.sort(key=lambda r: str(r["month"]))
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=BANK_PAYABLE_FIELDS)
        writer.writeheader()
        writer.writerows(normalized)


def upsert_bank_payable_month(path: Path, month: str, row: Dict[str, object]) -> None:
    rows = load_bank_payable_rows(path)
    by_month = {str(r["month"]): r for r in rows}
    by_month[month] = row
    save_bank_payable_rows(path, list(by_month.values()))


# -------------------------
# Rate and Schedule Engine
# -------------------------
def bank_rate_for_month(month: date, default_rate: float, rate_table: List[Tuple[date, float]]) -> float:
    # Use the latest known bank rate on or before the requested month.
    chosen = default_rate
    for eff, rate in rate_table:
        if eff <= month:
            chosen = rate
        else:
            break
    return chosen


def rate_override_for_month(month: date, rate_table: List[Tuple[date, float]]) -> bool:
    return any(eff <= month for eff, _ in rate_table)


def build_asset_schedule(
    asset: Asset,
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> List[Dict[str, object]]:
    """Build the month-by-month schedule for one asset.

    If a month already exists in the posted ledger, that month is reused as locked history.
    """
    balance = asset.lease_base
    rows: List[Dict[str, object]] = []
    posted_rows = posted_rows or {}

    for i in range(asset.payment_months):
        payment_due_date = add_months_keep_day(asset.start_date, i + 1)
        payment_due_key = payment_due_date.isoformat()
        posted_row = posted_rows.get((asset.asset_id, payment_due_key))

        if posted_row is not None:
            # Locked month: keep posted values and continue from posted ending balance.
            row = {
                "asset_id": asset.asset_id,
                "property_id": posted_row.get("property_id", asset.property_id),
                "allocation": posted_row.get("allocation", asset.allocation),
                "gl_account": posted_row.get("gl_account", asset.gl_account),
                "payment_month": payment_due_key,
                "due_date": str(posted_row.get("due_date", payment_due_key)),
                "billing_date": str(posted_row.get("billing_date", "")),
                "period": int(posted_row["period"]),
                "term_to_maturity": int(posted_row["term_to_maturity"]),
                "bank_rate_annual": round(float(posted_row["bank_rate_annual"]), 6),
                "nim_annual": round(float(posted_row["nim_annual"]), 6),
                "lease_rate_monthly": round(float(posted_row["lease_rate_monthly"]), 8),
                "lease_base": round(float(posted_row["lease_base"]), 2),
                "opening_balance": round(float(posted_row["opening_balance"]), 2),
                "payment": round(float(posted_row["payment"]), 2),
                "interest": round(float(posted_row["interest"]), 2),
                "amortization": round(float(posted_row["amortization"]), 2),
                "ending_balance": round(float(posted_row["ending_balance"]), 2),
                "rate_source": "posted_ledger",
            }
            rows.append(row)
            balance = float(posted_row["ending_balance"])
            if balance <= 0.0:
                break
            continue

        remaining = asset.payment_months - i
        # Open month: compute payment from current balance and remaining term.
        annual_bank_rate = bank_rate_for_month(month_start(payment_due_date), asset.bank_rate_annual, rate_table)
        annual_rate = annual_bank_rate + asset.nim_annual
        monthly_rate = effective_monthly_rate(annual_rate)
        payment = pmt(balance, monthly_rate, remaining)
        interest = balance * monthly_rate
        principal = payment - interest

        if principal > balance:
            principal = balance
            payment = interest + principal

        ending_balance = balance - principal
        if ending_balance <= 1e-8:
            ending_balance = 0.0

        rows.append(
            {
                "asset_id": asset.asset_id,
                "property_id": asset.property_id,
                "allocation": asset.allocation,
                "gl_account": asset.gl_account,
                "payment_month": payment_due_date.isoformat(),
                "due_date": payment_due_date.isoformat(),
                "billing_date": "",
                "period": i + 1,
                "term_to_maturity": remaining,
                "bank_rate_annual": round(annual_bank_rate, 6),
                "nim_annual": round(asset.nim_annual, 6),
                "lease_rate_monthly": round(monthly_rate, 8),
                "lease_base": round(asset.lease_base, 2),
                "opening_balance": round(balance, 2),
                "payment": round(payment, 2),
                "interest": round(interest, 2),
                "amortization": round(principal, 2),
                "ending_balance": round(ending_balance, 2),
                "rate_source": "rate_table"
                if rate_override_for_month(month_start(payment_due_date), rate_table)
                else "asset_default",
            }
        )

        balance = ending_balance
        if balance <= 0.0:
            break

    return rows


def build_schedule(
    assets: List[Asset],
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> List[Dict[str, object]]:
    """Build projected/actual schedule rows across all assets."""
    rows: List[Dict[str, object]] = []
    for asset in assets:
        rows.extend(build_asset_schedule(asset, rate_table, posted_rows=posted_rows))
    return rows


# ---------------------------------
# Snapshot and Invoice Transforming
# ---------------------------------
def amortize_until(
    asset: Asset,
    as_of: date,
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> Dict[str, float]:
    """Summarize payments, interest, amortization, and remaining balance up to as_of."""
    if as_of < asset.start_date:
        return {
            "installments_paid": 0,
            "interest_paid": 0.0,
            "amortization_paid": 0.0,
            "current_payment": 0.0,
            "balance": asset.lease_base,
            "months_remaining": asset.payment_months,
        }

    # Reuse schedule logic so snapshot math and invoice math stay aligned.
    schedule_rows = build_asset_schedule(asset, rate_table, posted_rows=posted_rows)
    paid_rows = [r for r in schedule_rows if parse_date(str(r["payment_month"])) <= as_of]

    periods = len(paid_rows)
    interest_paid = sum(float(r["interest"]) for r in paid_rows)
    amort_paid = sum(float(r["amortization"]) for r in paid_rows)
    current_payment = float(paid_rows[-1]["payment"]) if paid_rows else 0.0
    balance = float(paid_rows[-1]["ending_balance"]) if paid_rows else asset.lease_base

    months_remaining = max(0, asset.payment_months - periods)
    return {
        "installments_paid": periods,
        "interest_paid": interest_paid,
        "amortization_paid": amort_paid,
        "current_payment": current_payment,
        "balance": balance,
        "months_remaining": months_remaining,
    }


def build_snapshot(
    assets: List[Asset],
    rate_table: List[Tuple[date, float]],
    as_of: date,
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> List[Dict[str, object]]:
    """Return one current-state row per asset for reporting and exports."""
    rows: List[Dict[str, object]] = []
    for a in assets:
        m = amortize_until(a, as_of, rate_table, posted_rows=posted_rows)
        final_date = a.final_month
        days_to_maturity = (final_date - as_of).days
        active = a.status == "active" and m["balance"] > 0 and as_of >= a.start_date

        rows.append(
            {
                "asset_id": a.asset_id,
                "property_id": a.property_id,
                "allocation": a.allocation,
                "status": "active" if active else "inactive",
                "asset_value": round(a.asset_value, 2),
                "tax_amount": round(a.tax_amount, 2),
                "admin_expense": round(a.admin_expense, 2),
                "risk_cost_recovery": round(a.risk_cost_recovery, 2),
                "salvage_value": round(a.salvage_value, 2),
                "salvage_periods": a.salvage_periods,
                "lease_base": round(a.lease_base, 2),
                "start_date": a.start_date.isoformat(),
                "final_date": final_date.isoformat(),
                "term_months": a.term_months,
                "bank_rate_annual_default": round(a.bank_rate_annual, 6),
                "nim_annual": round(a.nim_annual, 6),
                "gl_account": a.gl_account,
                "installments_paid": int(m["installments_paid"]),
                "interest_paid": round(m["interest_paid"], 2),
                "amortization_paid": round(m["amortization_paid"], 2),
                "current_payment": round(m["current_payment"], 2),
                "balance": round(m["balance"], 2),
                "months_remaining": int(m["months_remaining"]),
                "days_to_maturity": days_to_maturity,
            }
        )
    return rows


def rows_total(rows: List[Dict[str, object]]) -> Dict[str, float]:
    total_assets = len(rows)
    active_assets = sum(1 for r in rows if r["status"] == "active")
    principal = sum(float(r["asset_value"]) for r in rows)
    lease_base = sum(float(r["lease_base"]) for r in rows)
    interest_paid = sum(float(r["interest_paid"]) for r in rows)
    amort_paid = sum(float(r["amortization_paid"]) for r in rows)
    balance = sum(float(r["balance"]) for r in rows)
    current_payment = sum(float(r["current_payment"]) for r in rows if r["status"] == "active")
    return {
        "total_assets": total_assets,
        "active_assets": active_assets,
        "asset_value_total": round(principal, 2),
        "lease_base_total": round(lease_base, 2),
        "interest_paid_total": round(interest_paid, 2),
        "amortization_paid_total": round(amort_paid, 2),
        "balance_total": round(balance, 2),
        "current_payment_active_total": round(current_payment, 2),
    }


def invoice_for_month(rows: List[Dict[str, object]], target_month: date) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    for r in rows:
        start_date = parse_date(str(r["start_date"]))
        final_date = parse_date(str(r["final_date"]))
        if month_start(start_date) <= target_month < month_start(final_date) and r["status"] == "active":
            out.append(
                {
                    "invoice_month": target_month.isoformat(),
                    "asset_id": r["asset_id"],
                    "property_id": r["property_id"],
                    "allocation": r["allocation"],
                    "gl_account": r["gl_account"],
                    "invoice_amount": r["current_payment"],
                }
            )
    return out


def invoice_for_month_from_schedule(
    rows: List[Dict[str, object]],
    target_month: date,
    billing_day: int,
) -> List[Dict[str, object]]:
    # Invoices are produced for installments due between previous and current billing runs.
    out: List[Dict[str, object]] = []
    previous_billing_date, billing_date = billing_cycle_window(target_month, billing_day)
    for r in rows:
        due_date = parse_date(str(r["payment_month"]))
        if not (previous_billing_date < due_date <= billing_date):
            continue

        out.append(
            {
                "invoice_month": target_month.isoformat(),
                "billing_date": billing_date.isoformat(),
                "due_date": due_date.isoformat(),
                "payment_month": r["payment_month"],
                "asset_id": r["asset_id"],
                "property_id": r["property_id"],
                "allocation": r["allocation"],
                "gl_account": r["gl_account"],
                "period": r["period"],
                "term_to_maturity": r["term_to_maturity"],
                "bank_rate_annual": r["bank_rate_annual"],
                "nim_annual": r["nim_annual"],
                "invoice_amount": r["payment"],
                "interest_amount": r["interest"],
                "amortization_amount": r["amortization"],
                "lease_rate_monthly": r["lease_rate_monthly"],
                "lease_base": r["lease_base"],
                "opening_balance": r["opening_balance"],
                "ending_balance": r["ending_balance"],
                "rate_source": r["rate_source"],
            }
        )
    return out


def write_csv(path: Path, rows: List[Dict[str, object]]) -> None:
    if not rows:
        return
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def safe_sheet_title(raw: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join(ch for ch in (raw or "") if ch not in invalid).strip()
    return (cleaned or "asset")[:31]


def asset_schedule_with_opening_row(
    asset: Asset,
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    opening_annual = bank_rate_for_month(month_start(asset.start_date), asset.bank_rate_annual, rate_table)
    opening_monthly = effective_monthly_rate(opening_annual + asset.nim_annual)
    rows.append(
        {
            "ledger_date": asset.start_date,
            "period": 0,
            "loan_rate_annual": opening_annual,
            "lease_rate_monthly": opening_monthly,
            "term_to_maturity": asset.payment_months,
            "installment": asset.lease_base,
            "interest": 0.0,
            "amortization": 0.0,
            "balance": asset.lease_base,
            "rate_source": "opening",
        }
    )

    schedule = build_asset_schedule(asset, rate_table, posted_rows=posted_rows)
    for row in schedule:
        rows.append(
            {
                "ledger_date": parse_date(str(row["payment_month"])),
                "period": int(row["period"]),
                "loan_rate_annual": float(row["bank_rate_annual"]),
                "lease_rate_monthly": float(row["lease_rate_monthly"]),
                "term_to_maturity": int(row["term_to_maturity"]),
                "installment": float(row["payment"]),
                "interest": float(row["interest"]),
                "amortization": float(row["amortization"]),
                "balance": float(row["ending_balance"]),
                "rate_source": str(row.get("rate_source", "")),
            }
        )
    return rows


def write_asset_one_pager_sheet(
    ws,
    asset: Asset,
    rate_table: List[Tuple[date, float]],
    lifecycle_status: str,
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    title_fill = PatternFill(fill_type="solid", fgColor="1F3A68")
    input_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF5")

    ws.merge_cells("A1:I1")
    ws["A1"] = "ALC - Acento Leasing Company Asset Calculation Model"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = title_fill

    input_rows = [
        ("Asset ID", asset.asset_id, None),
        ("Property ID", asset.property_id, None),
        ("Allocation", asset.allocation, None),
        ("Asset Description", asset.asset_description, None),
        ("Lifecycle Status", lifecycle_status, None),
        ("Asset Cost", asset.asset_value, "$#,##0.00"),
        ("Tax", asset.tax_amount, "$#,##0.00"),
        ("Admin Expenses", asset.admin_expense, "$#,##0.00"),
        ("Risk Cost Recovery", asset.risk_cost_recovery, "$#,##0.00"),
        ("Salvage Value", asset.salvage_value, "$#,##0.00"),
        ("Lease Base", asset.lease_base, "$#,##0.00"),
        ("Salvage Periods", asset.salvage_periods, None),
        ("Lifespan (months)", asset.term_months, None),
        ("NIM (annual)", asset.nim_annual, "0.00%"),
    ]
    for idx, (label, value, number_format) in enumerate(input_rows, start=3):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
        ws[f"A{idx}"].fill = input_fill
        ws[f"B{idx}"].fill = input_fill
        ws[f"A{idx}"].font = Font(bold=True)
        if number_format:
            ws[f"B{idx}"].number_format = number_format

    header_row = 3 + len(input_rows) + 2
    ws[f"A{header_row}"] = "Ledger Date"
    ws[f"B{header_row}"] = "Period"
    ws[f"C{header_row}"] = "Loan Rate"
    ws[f"D{header_row}"] = "Lease Rate"
    ws[f"E{header_row}"] = "Term to Maturity"
    ws[f"F{header_row}"] = "Installment"
    ws[f"G{header_row}"] = "Interest"
    ws[f"H{header_row}"] = "Amortization"
    ws[f"I{header_row}"] = "Balance"
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    schedule_rows = asset_schedule_with_opening_row(asset, rate_table, posted_rows=posted_rows)
    row_idx = header_row + 1
    for row in schedule_rows:
        ws[f"A{row_idx}"] = row["ledger_date"]
        ws[f"B{row_idx}"] = row["period"]
        ws[f"C{row_idx}"] = row["loan_rate_annual"]
        ws[f"D{row_idx}"] = row["lease_rate_monthly"]
        ws[f"E{row_idx}"] = row["term_to_maturity"]
        ws[f"F{row_idx}"] = row["installment"]
        ws[f"G{row_idx}"] = row["interest"]
        ws[f"H{row_idx}"] = row["amortization"]
        ws[f"I{row_idx}"] = row["balance"]
        row_idx += 1

    for r in range(header_row + 1, row_idx):
        ws[f"A{r}"].number_format = "yyyy-mm-dd"
        ws[f"C{r}"].number_format = "0.00%"
        ws[f"D{r}"].number_format = "0.00%"
        ws[f"F{r}"].number_format = "$#,##0.00"
        ws[f"G{r}"].number_format = "$#,##0.00"
        ws[f"H{r}"].number_format = "$#,##0.00"
        ws[f"I{r}"].number_format = "$#,##0.00"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.freeze_panes = f"A{header_row + 1}"


def write_inventory_sheet(
    ws,
    assets: List[Asset],
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]] = None,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    title_fill = PatternFill(fill_type="solid", fgColor="1F3A68")
    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF5")

    ws.merge_cells("A1:N1")
    ws["A1"] = "ALC - Asset Inventory"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = title_fill

    headers = [
        "Asset ID",
        "Property ID",
        "Allocation",
        "Description",
        "Status",
        "Start Date",
        "Final Date",
        "Term Months",
        "Asset Cost",
        "Lease Base",
        "Current Installment",
        "Outstanding Balance",
        "Bank APR",
        "NIM",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    snapshot_rows = build_snapshot(assets, rate_table, date.today(), posted_rows=posted_rows)
    asset_by_id = {asset.asset_id: asset for asset in assets}
    for row_idx, row in enumerate(snapshot_rows, start=4):
        asset = asset_by_id[str(row["asset_id"])]
        ws.cell(row=row_idx, column=1, value=row["asset_id"])
        ws.cell(row=row_idx, column=2, value=row["property_id"])
        ws.cell(row=row_idx, column=3, value=row["allocation"])
        ws.cell(row=row_idx, column=4, value=asset.asset_description)
        ws.cell(row=row_idx, column=5, value=row["status"])
        ws.cell(row=row_idx, column=6, value=parse_date(str(row["start_date"])))
        ws.cell(row=row_idx, column=7, value=parse_date(str(row["final_date"])))
        ws.cell(row=row_idx, column=8, value=row["term_months"])
        ws.cell(row=row_idx, column=9, value=row["asset_value"])
        ws.cell(row=row_idx, column=10, value=row["lease_base"])
        ws.cell(row=row_idx, column=11, value=row["current_payment"])
        ws.cell(row=row_idx, column=12, value=row["balance"])
        ws.cell(row=row_idx, column=13, value=row["bank_rate_annual_default"])
        ws.cell(row=row_idx, column=14, value=row["nim_annual"])

    for row_idx in range(4, len(snapshot_rows) + 4):
        ws.cell(row=row_idx, column=6).number_format = "yyyy-mm-dd"
        ws.cell(row=row_idx, column=7).number_format = "yyyy-mm-dd"
        ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=10).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=11).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=12).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=13).number_format = "0.00%"
        ws.cell(row=row_idx, column=14).number_format = "0.00%"

    widths = {
        "A": 12,
        "B": 12,
        "C": 14,
        "D": 28,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 14,
        "K": 18,
        "L": 18,
        "M": 10,
        "N": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def write_one_pager_workbook(
    assets: List[Asset],
    rate_table: List[Tuple[date, float]],
    posted_rows: Optional[Dict[Tuple[str, str], Dict[str, object]]],
    output_path: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for one-pager workbook generation") from exc

    ensure_parent(output_path)
    wb = Workbook()

    inventory_ws = wb.active
    inventory_ws.title = "Inventory"
    write_inventory_sheet(inventory_ws, assets, rate_table, posted_rows=posted_rows)

    snapshot_by_asset = {
        str(row["asset_id"]): row
        for row in build_snapshot(assets, rate_table, date.today(), posted_rows=posted_rows)
    }

    for asset in assets:
        desired = safe_sheet_title(asset.asset_id)
        ws = wb.create_sheet(desired)
        lifecycle_status = str(snapshot_by_asset.get(asset.asset_id, {}).get("status", "inactive"))
        write_asset_one_pager_sheet(ws, asset, rate_table, lifecycle_status, posted_rows=posted_rows)

    wb.save(output_path)


def print_totals(totals: Dict[str, float], title: str) -> None:
    print(f"\n{title}")
    print("-" * len(title))
    for k, v in totals.items():
        print(f"{k}: {v}")


# ----------------------
# Command Entry Handlers
# ----------------------
def run_snapshot(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    baseline_path = Path(args.baseline_config)
    input_paths = {
        "assets": assets_path,
        "rates": rates_path,
        "posted_ledger": posted_path,
        "baseline_config": baseline_path,
    }

    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_rows = posted_invoice_map(load_posted_invoices(posted_path))
    as_of = parse_date(args.as_of)
    rows = build_snapshot(assets, rates, as_of, posted_rows=posted_rows)
    totals = rows_total(rows)

    if args.output:
        write_csv(Path(args.output), rows)
        print(f"snapshot written to {args.output}")

    print_totals(totals, f"ALC Snapshot as of {as_of.isoformat()}")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "snapshot",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "as_of": as_of.isoformat(),
        "row_count": len(rows),
        "totals": totals,
        "input_hashes": collect_hashes(input_paths),
        "output_file": args.output or "",
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_invoice(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    closed_periods_path = Path(args.closed_periods)
    baseline_path = Path(args.baseline_config)
    backup_dir = Path(args.backup_dir)
    output_path = Path(args.output) if args.output else None
    input_paths = {
        "assets": assets_path,
        "rates": rates_path,
        "posted_ledger": posted_path,
        "closed_periods": closed_periods_path,
        "baseline_config": baseline_path,
    }
    backups: List[str] = []
    pre_hashes = collect_hashes(input_paths)

    baseline_cfg = load_baseline_config(baseline_path)
    target_month = month_start(parse_date(args.month + "-01"))
    go_live_str = str(baseline_cfg.get("go_live_date", "")).strip()
    if go_live_str:
        go_live_month = month_start(parse_date(go_live_str))
        if target_month < go_live_month and not args.allow_closed_adjustment:
            raise ValueError(
                f"invoice month {target_month.isoformat()} is before go-live {go_live_month.isoformat()}"
            )

    closed_months = load_closed_periods(closed_periods_path)
    target_month_key = month_key(target_month)
    if target_month_key in closed_months and not args.allow_closed_adjustment:
        raise ValueError(
            f"invoice month {target_month_key} is closed. Use --allow-closed-adjustment for authorized corrections."
        )

    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_existing = load_posted_invoices(posted_path)
    posted_rows = posted_invoice_map(posted_existing)
    schedule_rows = build_schedule(assets, rates, posted_rows=posted_rows)
    invoices = invoice_for_month_from_schedule(schedule_rows, target_month, args.billing_day)
    _, billing_date = billing_cycle_window(target_month, args.billing_day)

    total_invoice = round(sum(float(r["invoice_amount"]) for r in invoices), 2)
    print(f"invoice lines: {len(invoices)}")
    print(f"billing date ({target_month.isoformat()}): {billing_date.isoformat()}")
    print(f"invoice total ({target_month.isoformat()}): {total_invoice}")

    posted_batch = []
    timestamp = datetime.now().isoformat(timespec="seconds")
    # Persist the invoiced month as posted so later runs treat it as locked history.
    for row in invoices:
        posted_batch.append(
            {
                "asset_id": row["asset_id"],
                "property_id": row["property_id"],
                "allocation": row["allocation"],
                "gl_account": row["gl_account"],
                "payment_month": row["payment_month"],
                "due_date": row["due_date"],
                "billing_date": row["billing_date"],
                "period": row["period"],
                "term_to_maturity": row["term_to_maturity"],
                "bank_rate_annual": row["bank_rate_annual"],
                "nim_annual": row["nim_annual"],
                "lease_rate_monthly": row["lease_rate_monthly"],
                "lease_base": row["lease_base"],
                "opening_balance": row["opening_balance"],
                "payment": row["invoice_amount"],
                "interest": row["interest_amount"],
                "amortization": row["amortization_amount"],
                "ending_balance": row["ending_balance"],
                "rate_source": row["rate_source"],
                "posted_timestamp": timestamp,
            }
        )

    b = backup_file(posted_path, backup_dir, run_id)
    if b:
        backups.append(b)
    merged_posted = merge_posted_invoices(posted_existing, posted_batch)
    save_posted_invoices(posted_path, merged_posted)
    print(f"posted invoice ledger updated: {args.posted_ledger}")

    if args.close_period:
        b = backup_file(closed_periods_path, backup_dir, run_id)
        if b:
            backups.append(b)
        save_closed_periods(closed_periods_path, closed_months + [target_month_key])
        print(f"period closed: {target_month_key}")

    if args.output:
        if output_path is not None:
            b = backup_file(output_path, backup_dir, run_id)
            if b:
                backups.append(b)
            write_csv(output_path, invoices)
        print(f"invoice file written to {args.output}")

    one_pager_path = Path(args.one_pager_output)
    b = backup_file(one_pager_path, backup_dir, run_id)
    if b:
        backups.append(b)
    write_one_pager_workbook(assets, rates, posted_invoice_map(merged_posted), one_pager_path)
    print(f"one-pager workbook refreshed: {one_pager_path}")

    post_hashes = collect_hashes(input_paths)
    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "invoice",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "month": target_month.isoformat(),
        "billing_day": args.billing_day,
        "billing_date": billing_date.isoformat(),
        "allow_closed_adjustment": args.allow_closed_adjustment,
        "close_period": args.close_period,
        "invoice_lines": len(invoices),
        "invoice_total": total_invoice,
        "posted_ledger": str(posted_path),
        "input_hashes_before": pre_hashes,
        "input_hashes_after": post_hashes,
        "backups": backups,
        "output_file": args.output or "",
        "one_pager_output_file": str(one_pager_path),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_bank_payable(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    bank_payable_path = Path(args.bank_payable_file)
    backup_dir = Path(args.backup_dir)
    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_rows = posted_invoice_map(load_posted_invoices(posted_path))
    target_month = month_start(parse_date(args.month + "-01"))
    schedule_rows = build_schedule(assets, rates, posted_rows=posted_rows)
    invoices = invoice_for_month_from_schedule(schedule_rows, target_month, args.billing_day)
    _, billing_date = billing_cycle_window(target_month, args.billing_day)
    bank_payable = round(sum(float(r["invoice_amount"]) for r in invoices), 2)
    print(f"bank payable for {target_month.isoformat()} (billing date {billing_date.isoformat()}): {bank_payable}")

    backups: List[str] = []
    b = backup_file(bank_payable_path, backup_dir, run_id)
    if b:
        backups.append(b)

    month_str = month_key(target_month)
    upsert_bank_payable_month(
        bank_payable_path,
        month_str,
        {
            "month": month_str,
            "billing_date": billing_date.isoformat(),
            "invoice_lines": len(invoices),
            "bank_payable_total": bank_payable,
            "operator": args.operator,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        },
    )
    print(f"bank payable summary updated: {bank_payable_path}")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "bank-payable",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "month": target_month.isoformat(),
        "billing_day": args.billing_day,
        "billing_date": billing_date.isoformat(),
        "invoice_lines": len(invoices),
        "bank_payable_total": bank_payable,
        "bank_payable_file": str(bank_payable_path),
        "backups": backups,
        "input_hashes": collect_hashes(
            {
                "assets": assets_path,
                "rates": rates_path,
                "posted_ledger": posted_path,
                "baseline_config": Path(args.baseline_config),
            }
        ),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_close_period(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    closed_periods_path = Path(args.closed_periods)
    baseline_path = Path(args.baseline_config)
    backup_dir = Path(args.backup_dir)
    target_month = month_start(parse_date(args.month + "-01"))
    target_month_key = month_key(target_month)

    input_paths = {
        "closed_periods": closed_periods_path,
        "baseline_config": baseline_path,
    }
    pre_hashes = collect_hashes(input_paths)
    backups: List[str] = []

    baseline_cfg = load_baseline_config(baseline_path)
    go_live_str = str(baseline_cfg.get("go_live_date", "")).strip()
    if go_live_str:
        go_live_month = month_start(parse_date(go_live_str))
        if target_month < go_live_month:
            raise ValueError(
                f"close month {target_month.isoformat()} is before go-live {go_live_month.isoformat()}"
            )

    closed_months = load_closed_periods(closed_periods_path)
    already_closed = target_month_key in closed_months

    if already_closed:
        print(f"period already closed: {target_month_key}")
    else:
        b = backup_file(closed_periods_path, backup_dir, run_id)
        if b:
            backups.append(b)
        save_closed_periods(closed_periods_path, closed_months + [target_month_key])
        print(f"period closed: {target_month_key}")

    post_hashes = collect_hashes(input_paths)
    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "close-period",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "month": target_month.isoformat(),
        "already_closed": already_closed,
        "input_hashes_before": pre_hashes,
        "input_hashes_after": post_hashes,
        "backups": backups,
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_daily(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_rows = posted_invoice_map(load_posted_invoices(posted_path))
    as_of = parse_date(args.as_of)
    rows = build_snapshot(assets, rates, as_of, posted_rows=posted_rows)
    totals = rows_total(rows)
    print_totals(totals, f"Daily portfolio report ({as_of.isoformat()})")

    current_month = month_start(as_of)
    current_billing_date = billing_run_date_for_month(current_month, args.billing_day)
    if as_of <= current_billing_date:
        billing_month = current_month
        billing_date = current_billing_date
    else:
        billing_month = add_months(current_month, 1)
        billing_date = billing_run_date_for_month(billing_month, args.billing_day)

    schedule_rows = build_schedule(assets, rates, posted_rows=posted_rows)
    invoices = invoice_for_month_from_schedule(schedule_rows, billing_month, args.billing_day)
    due_count = len(invoices)
    due_amount = round(sum(float(r["invoice_amount"]) for r in invoices), 2)
    print(f"\nmonth_to_invoice: {billing_month.isoformat()}")
    print(f"billing_date: {billing_date.isoformat()}")
    print(f"invoice_lines_due: {due_count}")
    print(f"invoice_amount_due: {due_amount}")

    if as_of == billing_date - timedelta(days=1):
        print("\nREMINDER: Invoice run is due tomorrow.")
    elif as_of == billing_date:
        print("\nREMINDER: Invoice run is due today.")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "daily",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "as_of": as_of.isoformat(),
        "billing_day": args.billing_day,
        "billing_month": billing_month.isoformat(),
        "billing_date": billing_date.isoformat(),
        "portfolio_totals": totals,
        "invoice_lines_due": due_count,
        "invoice_amount_due": due_amount,
        "input_hashes": collect_hashes(
            {
                "assets": assets_path,
                "rates": rates_path,
                "posted_ledger": posted_path,
                "baseline_config": Path(args.baseline_config),
            }
        ),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_schedule(args: argparse.Namespace) -> None:
    # Useful for reviewing full projected/posted schedule before month-end close.
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_rows = posted_invoice_map(load_posted_invoices(posted_path))
    rows = build_schedule(assets, rates, posted_rows=posted_rows)

    if args.asset_id:
        rows = [r for r in rows if r["asset_id"] == args.asset_id]

    if args.output:
        write_csv(Path(args.output), rows)
        print(f"schedule written to {args.output}")

    print(f"schedule rows: {len(rows)}")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "schedule",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "asset_filter": args.asset_id or "",
        "row_count": len(rows),
        "output_file": args.output or "",
        "input_hashes": collect_hashes(
            {
                "assets": assets_path,
                "rates": rates_path,
                "posted_ledger": posted_path,
                "baseline_config": Path(args.baseline_config),
            }
        ),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_one_pager(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    output_path = Path(args.output)
    backup_dir = Path(args.backup_dir)

    assets = load_assets(assets_path)
    rates = load_rates(rates_path)
    posted_rows = posted_invoice_map(load_posted_invoices(posted_path))

    backups: List[str] = []
    b = backup_file(output_path, backup_dir, run_id)
    if b:
        backups.append(b)

    if args.asset_id:
        assets = [a for a in assets if a.asset_id == args.asset_id]

    if not assets:
        raise ValueError("no assets found for one-pager generation")

    write_one_pager_workbook(assets, rates, posted_rows, output_path)
    print(f"one-pager workbook written: {output_path}")
    print(f"tabs generated: {len(assets)}")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "one-pager",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "asset_filter": args.asset_id or "",
        "assets_processed": len(assets),
        "output_file": str(output_path),
        "backups": backups,
        "input_hashes": collect_hashes(
            {
                "assets": assets_path,
                "rates": rates_path,
                "posted_ledger": posted_path,
                "baseline_config": Path(args.baseline_config),
            }
        ),
        "output_hash": hash_file(output_path),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


def run_init_baseline(args: argparse.Namespace) -> None:
    run_id = generate_run_id()
    go_live = month_start(parse_date(args.go_live))

    assets_path = Path(args.assets)
    rates_path = Path(args.rates)
    posted_path = Path(args.posted_ledger)
    closed_periods_path = Path(args.closed_periods)
    baseline_path = Path(args.baseline_config)
    backup_dir = Path(args.backup_dir)

    inputs = {
        "assets": assets_path,
        "rates": rates_path,
        "posted_ledger": posted_path,
        "closed_periods": closed_periods_path,
        "baseline_config": baseline_path,
    }
    backups: List[str] = []
    for p in inputs.values():
        b = backup_file(p, backup_dir, run_id)
        if b:
            backups.append(b)

    if args.reset_posted_ledger:
        initialize_posted_ledger(posted_path)
        print(f"posted ledger initialized: {posted_path}")

    if args.reset_closed_periods:
        save_closed_periods(closed_periods_path, [])
        print(f"closed periods reset: {closed_periods_path}")

    baseline_cfg = {
        "go_live_date": go_live.isoformat(),
        "initialized_at": datetime.now().isoformat(timespec="seconds"),
        "initialized_by": args.operator,
        "run_id": run_id,
        "notes": args.notes or "",
        "script_version": SCRIPT_VERSION,
    }
    save_baseline_config(baseline_path, baseline_cfg)
    print(f"baseline config written: {baseline_path}")

    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "command": "init-baseline",
        "script_version": SCRIPT_VERSION,
        "operator": args.operator,
        "go_live_date": go_live.isoformat(),
        "reset_posted_ledger": args.reset_posted_ledger,
        "reset_closed_periods": args.reset_closed_periods,
        "backups": backups,
        "input_hashes_after": collect_hashes(inputs),
    }
    manifest_path = write_manifest(Path(args.manifest_dir), run_id, manifest)
    print(f"run manifest: {manifest_path}")


# ---------------------
# CLI Wiring and Main
# ---------------------
def build_parser() -> argparse.ArgumentParser:
    # All commands share the same source files to keep reporting and invoicing consistent.
    p = argparse.ArgumentParser(description="ALC lease tracker")
    sub = p.add_subparsers(dest="command", required=True)

    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--assets", default=str(DEFAULT_ASSETS), help="path to assets.csv")
    common.add_argument("--rates", default=str(DEFAULT_RATES), help="path to rates.csv")
    common.add_argument(
        "--posted-ledger", default=str(DEFAULT_POSTED_LEDGER), help="path to posted invoice ledger csv"
    )
    common.add_argument("--closed-periods", default=str(DEFAULT_CLOSED_PERIODS), help="path to closed periods csv")
    common.add_argument("--baseline-config", default=str(DEFAULT_BASELINE_CONFIG), help="path to baseline config json")
    common.add_argument(
        "--manifest-dir", default=str(DEFAULT_MANIFEST_DIR), help="directory for run manifest json files"
    )
    common.add_argument("--backup-dir", default=str(DEFAULT_BACKUP_DIR), help="directory for pre-write backups")
    common.add_argument("--operator", default="unknown", help="operator name or id for audit logs")
    common.add_argument("--billing-day", type=int, default=22, help="monthly invoice run day (shifted to next working day)")

    s1 = sub.add_parser("snapshot", parents=[common], help="point-in-time asset snapshot")
    s1.add_argument("--as-of", required=True, help="YYYY-MM-DD")
    s1.add_argument("--output", help="csv output path")
    s1.set_defaults(func=run_snapshot)

    s2 = sub.add_parser("invoice", parents=[common], help="invoice list for YYYY-MM")
    s2.add_argument("--month", required=True, help="YYYY-MM")
    s2.add_argument("--output", help="csv output path")
    s2.add_argument(
        "--one-pager-output",
        default=str(DEFAULT_ONE_PAGER),
        help="excel workbook path to refresh after invoice posting",
    )
    s2.add_argument(
        "--allow-closed-adjustment",
        action="store_true",
        help="allow invoicing a month already marked closed",
    )
    s2.add_argument("--close-period", action="store_true", help="mark the invoiced month as closed after posting")
    s2.set_defaults(func=run_invoice)

    s3 = sub.add_parser("bank-payable", parents=[common], help="bank payable for YYYY-MM")
    s3.add_argument("--month", required=True, help="YYYY-MM")
    s3.add_argument(
        "--bank-payable-file",
        default=str(DEFAULT_BANK_PAYABLE),
        help="csv file storing one bank-payable row per month",
    )
    s3.set_defaults(func=run_bank_payable)

    s8 = sub.add_parser("close-period", parents=[common], help="mark a month as closed without reposting invoices")
    s8.add_argument("--month", required=True, help="YYYY-MM")
    s8.set_defaults(func=run_close_period)

    s4 = sub.add_parser("daily", parents=[common], help="daily operating report")
    s4.add_argument("--as-of", required=True, help="YYYY-MM-DD")
    s4.set_defaults(func=run_daily)

    s5 = sub.add_parser("schedule", parents=[common], help="projected amortization schedule")
    s5.add_argument("--asset-id", help="limit schedule to one asset")
    s5.add_argument("--output", help="csv output path")
    s5.set_defaults(func=run_schedule)

    s6 = sub.add_parser("init-baseline", parents=[common], help="initialize clean production baseline")
    s6.add_argument("--go-live", required=True, help="go-live date YYYY-MM-DD (e.g. 2026-07-01)")
    s6.add_argument("--reset-posted-ledger", action="store_true", help="reset posted ledger to header only")
    s6.add_argument("--reset-closed-periods", action="store_true", help="reset closed periods file to empty")
    s6.add_argument("--notes", help="optional baseline notes")
    s6.set_defaults(func=run_init_baseline)

    s7 = sub.add_parser("one-pager", parents=[common], help="create/update one worksheet tab per asset")
    s7.add_argument("--asset-id", help="generate workbook for one asset only")
    s7.add_argument(
        "--output",
        default=str(DEFAULT_ONE_PAGER),
        help="excel workbook output path (created once, tabs updated by asset)",
    )
    s7.set_defaults(func=run_one_pager)

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()